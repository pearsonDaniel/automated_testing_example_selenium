# login_page.py
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from locators.login_locators import LoginPageLocators
from src.pages.base_page import BasePage
import openpyxl
import time
import os

spreadsheet = openpyxl.load_workbook(r"test\test_resources\user_credentials.xlsx")
credentials = spreadsheet.active
username = credentials.cell(row=2, column=1).value
password = credentials.cell(row=2, column=2).value

class LoginPage(BasePage):
    def __init__(self, driver):
        self.driver = driver

    def dismiss_chrome_alert_by_enter(self):
        """
        Sends Enter to the browser window to dismiss Chrome native alerts.
        Safe even if no alert is present.
        """
        try:
            self.driver.switch_to.active_element.send_keys(Keys.ENTER)
            return True
        except Exception:
            return False

    def login(self):
        self.driver.maximize_window()

        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(LoginPageLocators.USERNAME_FIELD)
        ).send_keys(username)

        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(LoginPageLocators.PASSWORD_FIELD)
        ).send_keys(password)

        WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(LoginPageLocators.LOGIN_BUTTON)
        ).click()

        # Try pressing Enter a few times in case the alert appears slightly delayed
        for _ in range(5):
            self.dismiss_chrome_alert_by_enter()
            time.sleep(0.3)


    def login_env(self):
        self.driver.maximize_window()
        username = os.getenv("LOGIN_USERNAME")
        password = os.getenv("LOGIN_PASSWORD")

        if not username or not password:
            raise ValueError(
                "Missing required environment variables LOGIN_USERNAME and LOGIN_PASSWORD."
            )
        
        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(LoginPageLocators.USERNAME_FIELD)
        ).send_keys(username)

        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(LoginPageLocators.PASSWORD_FIELD)
        ).send_keys(password)

        WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(LoginPageLocators.LOGIN_BUTTON)
        ).click()

        # Try pressing Enter a few times in case the alert appears slightly delayed
        for _ in range(5):
            self.dismiss_chrome_alert_by_enter()
            time.sleep(0.3)